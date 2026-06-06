import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Segmentador de Fases — SitStand", page_icon="📊", layout="wide")

st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #f8f9fa; }
[data-testid="stSidebar"] { background: #ffffff; border-right: 1px solid #dee2e6; }
.block-container { padding-top: 1.5rem; }
</style>
""", unsafe_allow_html=True)

CHART_SIZE        = 580   # px — quadrado
PHASE_COLORS      = ["rgba(220,53,69,0.10)","rgba(40,167,69,0.10)","rgba(0,123,255,0.10)"]
PHASE_LINE_COLORS = ["#dc3545","#28a745","#0d6efd"]
PHASE_NAMES       = ["P1","P2","P3"]

# ── Session state ──────────────────────────────────────────────────────────────
for k, v in [("people",[]),("summary",{}),("divisions",{}),("excluded",set()),
             ("grp_d1",None),("grp_d2",None)]:
    if k not in st.session_state:
        st.session_state[k] = v

# ── Carregar arquivo ───────────────────────────────────────────────────────────
def load_file(uploaded):
    wb = openpyxl.load_workbook(uploaded)
    summary = {}
    for sname in wb.sheetnames:
        if sname.strip().lower() == "resumo":
            for r in list(wb[sname].iter_rows(values_only=True))[1:]:
                if r[0] and r[2]:
                    try: summary[str(r[0]).strip()] = {"n_ciclos":r[1],"duracao_ms":float(r[2])}
                    except: pass
            break
    people = []
    for sname in wb.sheetnames:
        if sname.strip().lower() == "resumo": continue
        ws  = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r[0] is not None and r[1] is not None]
        if not data_rows: continue
        df = pd.DataFrame(data_rows, columns=["fase_norm","media","dp","upper","lower"])
        df = df.apply(pd.to_numeric, errors="coerce").dropna(subset=["fase_norm","media"])
        name = sname.strip()
        dur_ms   = next((v["duracao_ms"] for k,v in summary.items() if k in name or name in k), None)
        n_ciclos = next((v["n_ciclos"]   for k,v in summary.items() if k in name or name in k), None)
        people.append({"name":name,"data":df,"dur_ms":dur_ms,"n_ciclos":n_ciclos})
    return people, summary

# ── Métricas por fase ─────────────────────────────────────────────────────────
def calc_metrics(df, d1, d2, dur_ms):
    x_min = float(df["fase_norm"].min())
    x_max = float(df["fase_norm"].max())
    bounds = [x_min, d1, d2, x_max]
    total_range = x_max - x_min
    dur_s = dur_ms/1000 if dur_ms else None
    out = {}
    for i, p in enumerate(PHASE_NAMES):
        seg = df[(df["fase_norm"] >= bounds[i]) & (df["fase_norm"] <= bounds[i+1])]
        if seg.empty:
            for k in ["dur_s","max_acc","min_acc","range_acc","auc_pos","auc_neg"]:
                out[f"{k}_{p}"] = np.nan
            continue
        phase_frac = (bounds[i+1]-bounds[i]) / total_range
        out[f"dur_s_{p}"]    = round(phase_frac*dur_s*1000, 1) if dur_s else np.nan  # em ms
        y = seg["media"].values
        t = (seg["fase_norm"].values-bounds[i])/(bounds[i+1]-bounds[i]) * phase_frac * (dur_s or 1)
        out[f"max_acc_{p}"]   = round(float(np.max(y)),4)
        out[f"min_acc_{p}"]   = round(float(np.min(y)),4)
        out[f"range_acc_{p}"] = round(float(np.max(y)-np.min(y)),4)
        out[f"auc_pos_{p}"]   = round(float(np.trapezoid(np.where(y>0,y,0), t)),4)
        out[f"auc_neg_{p}"]   = round(float(np.trapezoid(np.where(y<0,y,0), t)),4)
    return out

# ── Resultante do grupo ────────────────────────────────────────────────────────
def compute_group(people, excluded):
    active = [p for p in people if p["name"] not in excluded]
    if len(active) < 2: return None
    x_common = np.linspace(0,1,300)
    matrix = []
    for p in active:
        df = p["data"]
        x = df["fase_norm"].values
        x_norm = (x-x.min())/(x.max()-x.min())
        matrix.append(np.interp(x_common, x_norm, df["media"].values))
    mat  = np.array(matrix)
    mean = np.mean(mat, axis=0)
    sd   = np.std(mat, axis=0, ddof=1)
    return pd.DataFrame({"fase_norm":x_common,"media":mean,"dp":sd,"upper":mean+sd,"lower":mean-sd})

# ── Figura quadrada ────────────────────────────────────────────────────────────
def build_fig(df, d1, d2, title=""):
    x_min = float(df["fase_norm"].min())
    x_max = float(df["fase_norm"].max())
    bounds = [x_min, d1, d2, x_max]
    fig = go.Figure()

    # Banda ±DP
    fig.add_trace(go.Scatter(
        x=list(df["fase_norm"])+list(df["fase_norm"])[::-1],
        y=list(df["upper"])+list(df["lower"])[::-1],
        fill="toself", fillcolor="rgba(150,150,150,0.15)",
        line=dict(color="rgba(0,0,0,0)"), hoverinfo="skip", showlegend=False))
    for col in ["upper","lower"]:
        fig.add_trace(go.Scatter(x=df["fase_norm"], y=df[col], mode="lines",
            line=dict(color="rgba(130,130,130,0.5)",width=1,dash="dot"),
            showlegend=False, hoverinfo="skip"))

    # Fundos das fases
    for i in range(3):
        fig.add_shape(type="rect", x0=bounds[i], x1=bounds[i+1], y0=0, y1=1, yref="paper",
            fillcolor=PHASE_COLORS[i], line=dict(width=0), layer="below")
        fig.add_annotation(x=(bounds[i]+bounds[i+1])/2, y=0.97, yref="paper",
            text=f"<b>{PHASE_NAMES[i]}</b>", showarrow=False,
            font=dict(color=PHASE_LINE_COLORS[i],size=14), yanchor="top", xanchor="center")

    # Linhas divisórias
    labels = ["Início P2", "Início P3"]
    for i, m in enumerate([d1, d2]):
        fig.add_vline(x=m, line=dict(color=PHASE_LINE_COLORS[i],width=2,dash="dash"),
            annotation_text=f"{labels[i]}={m:.3f}",
            annotation_font=dict(color=PHASE_LINE_COLORS[i],size=11),
            annotation_position="top right" if i==0 else "top left")

    # Curva média
    fig.add_trace(go.Scatter(x=df["fase_norm"], y=df["media"], mode="lines",
        line=dict(color="#212529",width=2.5), name="Média",
        hovertemplate="fase=%{x:.3f}<br>%{y:.3f} m/s²<extra></extra>"))

    fig.update_layout(
        title=dict(text=title, font=dict(color="#1a202c",size=13), x=0),
        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
        width=CHART_SIZE, height=CHART_SIZE,
        margin=dict(l=60, r=20, t=50, b=55),
        showlegend=False,
        xaxis=dict(title="Fase normalizada", color="#495057",
                   gridcolor="#e9ecef", zerolinecolor="#adb5bd"),
        yaxis=dict(title="Aceleração (m/s²)", color="#495057",
                   gridcolor="#e9ecef", zerolinecolor="#adb5bd"),
    )
    return fig

# ── Tabela de métricas ─────────────────────────────────────────────────────────
def build_metrics_df(people, excluded, divisions):
    rows = []
    for p in people:
        name = p["name"]
        if name in excluded: continue
        div = divisions.get(name)
        if not div: continue
        m = calc_metrics(p["data"], div["d1"], div["d2"], p.get("dur_ms"))
        short = name.split("sitstand")[0].split("SitStand")[0].strip().rstrip("-").strip()
        row = {
            "Pessoa": short,
            "N ciclos": p.get("n_ciclos",""),
            "Dur. média ciclo (ms)": p.get("dur_ms",""),
            "Início P2 (fase_norm)": round(div["d1"],4),
            "Início P3 (fase_norm)": round(div["d2"],4),
        }
        # Organizado por métrica, não por fase
        for pn in PHASE_NAMES:
            row[f"Duração_{pn} (ms)"] = m.get(f"dur_s_{pn}","")
        for pn in PHASE_NAMES:
            row[f"MaxAcc_{pn} (m/s²)"] = m.get(f"max_acc_{pn}","")
        for pn in PHASE_NAMES:
            row[f"MinAcc_{pn} (m/s²)"] = m.get(f"min_acc_{pn}","")
        for pn in PHASE_NAMES:
            row[f"Range_{pn} (m/s²)"] = m.get(f"range_acc_{pn}","")
        for pn in PHASE_NAMES:
            row[f"AUC+_{pn} (m/s²·s)"] = m.get(f"auc_pos_{pn}","")
        for pn in PHASE_NAMES:
            row[f"AUC-_{pn} (m/s²·s)"] = m.get(f"auc_neg_{pn}","")
        rows.append(row)
    return pd.DataFrame(rows)

def add_stats_rows(df):
    """Adiciona linhas de estatísticas do grupo ao final da tabela."""
    if df.empty: return df, []
    num_cols = df.select_dtypes(include="number").columns.tolist()
    stats = {
        "Média":    df[num_cols].mean(),
        "Mediana":  df[num_cols].median(),
        "Q1 (25%)": df[num_cols].quantile(0.25),
        "Q3 (75%)": df[num_cols].quantile(0.75),
        "DP":       df[num_cols].std(ddof=1),
    }
    stat_rows = []
    for label, vals in stats.items():
        row = {c: "" for c in df.columns}
        row["Pessoa"] = label
        for c in num_cols:
            v = vals[c]
            row[c] = round(v,4) if pd.notna(v) else ""
        stat_rows.append(row)
    stat_df = pd.DataFrame(stat_rows)
    combined = pd.concat([df, stat_df], ignore_index=True)
    return combined, list(range(len(df), len(df)+len(stat_rows)))

# ── Exportar Excel ─────────────────────────────────────────────────────────────
def make_excel(people, excluded, divisions, grp_df, grp_d1, grp_d2):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # Aba 1: Métricas individuais + estatísticas
        df_ind = build_metrics_df(people, excluded, divisions)
        if not df_ind.empty:
            combined, stat_idx = add_stats_rows(df_ind)
            combined.to_excel(writer, sheet_name="Métricas Individuais", index=False)
            ws = writer.sheets["Métricas Individuais"]
            # Cabeçalho
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="2C3E50")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.row_dimensions[1].height = 30
            # Linhas de estatísticas em negrito + fundo suave
            for si in stat_idx:
                row_num = si + 2  # +1 header +1 excel 1-indexed
                fill = PatternFill("solid", fgColor="D5E8D4")
                for cell in ws[row_num]:
                    cell.font = Font(bold=True)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center")
            # Larguras
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = 18
            ws.freeze_panes = "B2"

        # Aba 2: Curva da resultante do grupo
        if grp_df is not None:
            grp_df.rename(columns={
                "fase_norm":"Fase_norm","media":"Média (m/s²)","dp":"DP (m/s²)",
                "upper":"Média+DP","lower":"Média-DP"
            }).to_excel(writer, sheet_name="Resultante Grupo", index=False)
            ws2 = writer.sheets["Resultante Grupo"]
            for cell in ws2[1]:
                cell.fill = PatternFill("solid", fgColor="1A5276")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for col in ws2.columns:
                ws2.column_dimensions[get_column_letter(col[0].column)].width = 16

        # Aba 3: Métricas da resultante grupo
        if grp_df is not None and grp_d1 and grp_d2:
            gm = calc_metrics(grp_df, grp_d1, grp_d2, None)
            n_ativos = len([p for p in people if p["name"] not in excluded])
            grp_row = {
                "Grupo": f"Resultante Grupo (n={n_ativos})",
                "Início P2 (fase_norm)": round(grp_d1,4),
                "Início P3 (fase_norm)": round(grp_d2,4),
            }
            for pn in PHASE_NAMES:
                grp_row[f"MaxAcc_{pn} (m/s²)"]  = gm.get(f"max_acc_{pn}","")
            for pn in PHASE_NAMES:
                grp_row[f"MinAcc_{pn} (m/s²)"]  = gm.get(f"min_acc_{pn}","")
            for pn in PHASE_NAMES:
                grp_row[f"Range_{pn} (m/s²)"]   = gm.get(f"range_acc_{pn}","")
            for pn in PHASE_NAMES:
                grp_row[f"AUC+_{pn} (m/s²·s)"]  = gm.get(f"auc_pos_{pn}","")
            for pn in PHASE_NAMES:
                grp_row[f"AUC-_{pn} (m/s²·s)"]  = gm.get(f"auc_neg_{pn}","")
            pd.DataFrame([grp_row]).to_excel(writer, sheet_name="Métricas Grupo", index=False)
            ws3 = writer.sheets["Métricas Grupo"]
            for cell in ws3[1]:
                cell.fill = PatternFill("solid", fgColor="117A65")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for col in ws3.columns:
                ws3.column_dimensions[get_column_letter(col[0].column)].width = 20

    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### 📂 Arquivo")
    uploaded = st.file_uploader("Selecione o .xlsx", type=["xlsx","xls"])
    if uploaded:
        if st.button("▶ Carregar / Recarregar"):
            people, summary = load_file(uploaded)
            st.session_state.people    = people
            st.session_state.summary   = summary
            st.session_state.divisions = {}
            st.session_state.excluded  = set()
            st.session_state.grp_d1    = None
            st.session_state.grp_d2    = None
            st.rerun()
    st.divider()
    if st.session_state.people:
        excl    = st.session_state.excluded
        ativos  = [p for p in st.session_state.people if p["name"] not in excl]
        marcados = sum(1 for p in ativos if p["name"] in st.session_state.divisions)
        st.metric("Pessoas ativas", len(ativos))
        st.metric("Marcados", f"{marcados} / {len(ativos)}")
        st.divider()
        st.markdown("**Como usar:** Arraste **D1** e **D2** para dividir P1|P2 e P2|P3. Marque **Excluir** para remover.")
        st.divider()
        grp_df = compute_group(st.session_state.people, st.session_state.excluded)
        gd1 = st.session_state.grp_d1 or 0.33
        gd2 = st.session_state.grp_d2 or 0.67
        excel_data = make_excel(
            st.session_state.people, st.session_state.excluded,
            st.session_state.divisions, grp_df, gd1, gd2)
        st.download_button("⬇ Exportar Excel (.xlsx)", data=excel_data,
            file_name="metricas_sitstand.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
st.title("📊 Segmentador de Fases — SitStand")

if not st.session_state.people:
    st.info("👈 Carregue o arquivo **.xlsx** na barra lateral e clique em **Carregar**.")
    st.stop()

# ── Gráficos individuais ───────────────────────────────────────────────────────
st.subheader("Indivíduos")
people = st.session_state.people

for row_start in range(0, len(people), 2):
    cols = st.columns(2, gap="large")
    for col_idx, person in enumerate(people[row_start: row_start+2]):
        name   = person["name"]
        df     = person["data"]
        dur_ms = person.get("dur_ms")
        is_excl = name in st.session_state.excluded

        x_min = float(df["fase_norm"].min())
        x_max = float(df["fase_norm"].max())
        x_rng = x_max - x_min
        step  = round(x_rng/300, 5)
        saved = st.session_state.divisions.get(name, {})
        def_d1 = saved.get("d1", round(x_min+x_rng*0.33, 5))
        def_d2 = saved.get("d2", round(x_min+x_rng*0.67, 5))
        short  = name.split("sitstand")[0].split("SitStand")[0].strip().rstrip("-").strip()

        with cols[col_idx]:
            hc = st.columns([5,1])
            with hc[1]:
                excl_val = st.checkbox("Excluir", value=is_excl, key=f"excl_{name}")
                if excl_val: st.session_state.excluded.add(name)
                else:        st.session_state.excluded.discard(name)
            if excl_val:
                hc[0].markdown(f"~~**{short}**~~ ❌ *excluída*")
                st.divider(); continue

            sc = st.columns(2)
            with sc[0]:
                d1 = st.slider("🔴 Início P2", float(x_min), float(x_max-step*2),
                    float(def_d1), float(step), format="%.3f", key=f"d1_{name}")
            with sc[1]:
                d2 = st.slider("🟢 Início P3", float(d1+step), float(x_max),
                    float(max(def_d2, d1+step*2)), float(step), format="%.3f", key=f"d2_{name}")
            st.session_state.divisions[name] = {"d1":d1,"d2":d2}

            fig = build_fig(df, d1, d2, title=short)
            st.plotly_chart(fig, use_container_width=False, key=f"chart_{name}")

            # Métricas inline
            m = calc_metrics(df, d1, d2, dur_ms)
            mc = st.columns(3)
            for i, pn in enumerate(PHASE_NAMES):
                dur_v = m.get(f"dur_s_{pn}")
                with mc[i]:
                    dur_txt = f"({dur_v:.0f}ms)" if dur_v and not np.isnan(float(dur_v)) else ""
                    st.markdown(f"**{pn}** {dur_txt}")
                    st.markdown(f"""<small>
Max: **{m.get(f'max_acc_{pn}','—')}**<br>
Min: **{m.get(f'min_acc_{pn}','—')}**<br>
Range: **{m.get(f'range_acc_{pn}','—')}**<br>
AUC+: **{m.get(f'auc_pos_{pn}','—')}**<br>
AUC−: **{m.get(f'auc_neg_{pn}','—')}**
</small>""", unsafe_allow_html=True)
            st.divider()

# ── Tabela de métricas + estatísticas do grupo ─────────────────────────────────
st.subheader("📋 Tabela de métricas — prévia do Excel")
df_tab = build_metrics_df(st.session_state.people, st.session_state.excluded, st.session_state.divisions)

if df_tab.empty:
    st.info("Marque as divisões nos gráficos acima para ver as métricas aqui.")
else:
    combined, stat_idx = add_stats_rows(df_tab)

    # Estilizar: negrito nas linhas de estatísticas + fundo verde claro
    def style_rows(row):
        if row.name in stat_idx:
            return ["font-weight: bold; background-color: #d5e8d4"] * len(row)
        return [""] * len(row)

    styled = combined.style.apply(style_rows, axis=1).format(
        {c: "{:.4f}" for c in combined.select_dtypes("number").columns}, na_rep=""
    )
    st.dataframe(styled, use_container_width=True, hide_index=True)
    st.caption("🟢 Linhas em verde = estatísticas do grupo | Aceleração em m/s² | AUC em m/s²·s | Duração em ms")

    # Botão de exportar a tabela
    def make_table_excel(combined, stat_idx):
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            combined.to_excel(writer, sheet_name="Métricas Individuais", index=False)
            ws = writer.sheets["Métricas Individuais"]
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="2C3E50")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.row_dimensions[1].height = 30
            for si in stat_idx:
                for cell in ws[si + 2]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D5E8D4")
                    cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = 18
            ws.freeze_panes = "B2"
        return buf.getvalue()

    st.download_button(
        "⬇ Exportar esta tabela (.xlsx)",
        data=make_table_excel(combined, stat_idx),
        file_name="metricas_individuais.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.divider()

# ── Resultante do grupo (ao final) ─────────────────────────────────────────────
grp_df = compute_group(st.session_state.people, st.session_state.excluded)

st.subheader("📈 Resultante do Grupo")
if grp_df is None:
    st.warning("Nenhuma pessoa ativa suficiente para calcular a resultante.")
else:
    sc1, sc2, _ = st.columns([1,1,2])
    with sc1:
        gd1 = st.slider("🔴 Início P2 — grupo", 0.0, 0.97,
            float(st.session_state.grp_d1 or 0.33), 0.001, format="%.3f", key="grp_d1_s")
    with sc2:
        gd2 = st.slider("🟢 Início P3 — grupo", float(gd1+0.001), 1.0,
            float(max(st.session_state.grp_d2 or 0.67, gd1+0.01)), 0.001, format="%.3f", key="grp_d2_s")
    st.session_state.grp_d1 = gd1
    st.session_state.grp_d2 = gd2

    n_ativos = len([p for p in st.session_state.people if p["name"] not in st.session_state.excluded])
    gcol, _ = st.columns([1,1])
    with gcol:
        fig_grp = build_fig(grp_df, gd1, gd2, title=f"Resultante do Grupo — n={n_ativos}")
        st.plotly_chart(fig_grp, use_container_width=False, key="chart_grupo")

    gm = calc_metrics(grp_df, gd1, gd2, None)
    mc = st.columns(3)
    for i, pn in enumerate(PHASE_NAMES):
        with mc[i]:
            st.markdown(f"**{pn}**")
            st.markdown(f"""<small>
Max: **{gm.get(f'max_acc_{pn}','—')} m/s²**<br>
Min: **{gm.get(f'min_acc_{pn}','—')} m/s²**<br>
Range: **{gm.get(f'range_acc_{pn}','—')} m/s²**<br>
AUC+: **{gm.get(f'auc_pos_{pn}','—')} m/s²·s**<br>
AUC−: **{gm.get(f'auc_neg_{pn}','—')} m/s²·s**
</small>""", unsafe_allow_html=True)

    # ── Export dedicado da resultante ──────────────────────────────────────────
    st.markdown("---")
    def make_grp_excel(grp_df, gd1, gd2, n_ativos):
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            # Aba 1: Matriz da curva (fase_norm, Média, DP, +DP, -DP)
            df_curva = grp_df.rename(columns={
                "fase_norm": "Fase_norm",
                "media":     "Média (m/s²)",
                "dp":        "DP (m/s²)",
                "upper":     "Média+DP (m/s²)",
                "lower":     "Média-DP (m/s²)",
            })
            df_curva.to_excel(writer, sheet_name="Curva Resultante", index=False)
            ws = writer.sheets["Curva Resultante"]
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="1A5276")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = 18

            # Aba 2: Métricas por fase da resultante
            gm = calc_metrics(grp_df, gd1, gd2, None)
            grp_row = {
                "Grupo": f"Resultante Grupo (n={n_ativos})",
                "Início P2 (fase_norm)": round(gd1, 4),
                "Início P3 (fase_norm)": round(gd2, 4),
            }
            for pn in PHASE_NAMES:
                grp_row[f"MaxAcc_{pn} (m/s²)"]  = gm.get(f"max_acc_{pn}", "")
            for pn in PHASE_NAMES:
                grp_row[f"MinAcc_{pn} (m/s²)"]  = gm.get(f"min_acc_{pn}", "")
            for pn in PHASE_NAMES:
                grp_row[f"Range_{pn} (m/s²)"]   = gm.get(f"range_acc_{pn}", "")
            for pn in PHASE_NAMES:
                grp_row[f"AUC+_{pn} (m/s²·s)"]  = gm.get(f"auc_pos_{pn}", "")
            for pn in PHASE_NAMES:
                grp_row[f"AUC-_{pn} (m/s²·s)"]  = gm.get(f"auc_neg_{pn}", "")
            pd.DataFrame([grp_row]).to_excel(writer, sheet_name="Métricas Resultante", index=False)
            ws2 = writer.sheets["Métricas Resultante"]
            for cell in ws2[1]:
                cell.fill = PatternFill("solid", fgColor="117A65")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for col in ws2.columns:
                ws2.column_dimensions[get_column_letter(col[0].column)].width = 20

        return buf.getvalue()

    grp_excel = make_grp_excel(grp_df, gd1, gd2, n_ativos)
    st.download_button(
        "⬇ Exportar Resultante do Grupo (.xlsx)",
        data=grp_excel,
        file_name="resultante_grupo.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption("2 abas: **Curva Resultante** (fase_norm, Média, DP, ±DP) e **Métricas Resultante** por fase")
